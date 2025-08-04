import streamlit as st
import pandas as pd
import numpy as np
import calendar
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# --- App config ---
st.set_page_config(page_title="Monthly Report Processor", layout="wide")

# --- Credentials ---
USERNAME = "admin"
PASSWORD = "report123"

# --- Auth & Login ---
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("🔐 Login Required")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    login_button = st.button("Login")
    if login_button:
        if username == USERNAME and password == PASSWORD:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("❌ Incorrect username or password")
    st.stop()

# --- UI Header ---
st.title("📊 Monthly Report Generator")
uploaded_files = st.file_uploader(
    "Upload multiple Excel files (YYYYMM.xlsx format)", type="xlsx", accept_multiple_files=True
)

# --- Helper: Format and download Excel ---
def format_and_download(df_result, report_type="official"):
    buffer = io.BytesIO()
    temp_file_name = "temp_report.xlsx"
    with pd.ExcelWriter(temp_file_name, engine='openpyxl') as writer:
        df_result.to_excel(writer, index=False, sheet_name="Report")

    wb = load_workbook(temp_file_name)
    ws = wb["Report"]
    col_index_map = {col: idx + 1 for idx, col in enumerate(df_result.columns)}
    month_names = [calendar.month_abbr[i] for i in range(1, 13)]
    number_cols = month_names + ['Grand Total']

    for col in number_cols:
        col_idx = col_index_map.get(col)
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter][1:]:
            cell.number_format = '#,##0_);[Red](#,##0)'

    if 'Customer' in col_index_map:
        ws.column_dimensions[get_column_letter(col_index_map['Customer'])].width = 35
    if 'Item Detail' in col_index_map:
        ws.column_dimensions[get_column_letter(col_index_map['Item Detail'])].width = 35
    if 'Year' in col_index_map:
        col_letter = get_column_letter(col_index_map['Year'])
        for cell in ws[col_letter][1:]:
            cell.alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{report_type}_monthly_report.xlsx"
    st.download_button(
        label=f"📥 Download {report_type.capitalize()} Monthly Report",
        data=buffer,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --- Report 1: Official ---
def generate_official_report(files):
    df_list = []
    for i, file in enumerate(files):
        filename = file.name
        month = filename[4:6]
        year = filename[0:4]

        header = pd.read_excel(file, nrows=3, header=None)
        last_col_index = header.iloc[2].tolist().index("Total") if "Total" in header.iloc[2].tolist() else None
        usecols = list(range(last_col_index)) if last_col_index is not None else None

        df = pd.read_excel(file, skiprows=2, nrows=49, usecols=usecols)
        df.columns.values[0:3] = ['Type', 'Item', 'Item Detail']
        df_long = df.melt(id_vars=['Type', 'Item', 'Item Detail'], var_name='Site', value_name='Amount')
        df_long['Month'] = month
        df_long['Year'] = year
        df_list.append(df_long)

    df_final = pd.concat(df_list, ignore_index=True)
    df_final['Amount'] = pd.to_numeric(df_final['Amount'], errors='coerce').fillna(0)
    df_final['Month'] = df_final['Month'].astype(str).str.zfill(2)
    month_map = {f"{i:02d}": calendar.month_abbr[i] for i in range(1, 13)}

    pivot_df = pd.pivot_table(
        df_final,
        index=['Site', 'Item Detail', 'Year'],
        columns='Month',
        values='Amount',
        aggfunc='sum',
        fill_value=0,
        observed=False
    )
    pivot_df['Grand Total'] = pivot_df.sum(axis=1)
    pivot_df = pivot_df.reset_index()
    pivot_df = pivot_df.rename(columns=month_map)

    final_columns = ['Site', 'Item Detail', 'Year'] + list(month_map.values()) + ['Grand Total']
    pivot_df = pivot_df.reindex(columns=final_columns)
    return pivot_df

# --- Report 2: Customer ---
def generate_customer_report(files):
    warehouse_site_map = {
        "WAREHOUSE ( ACW )": "ACW", "WAREHOUSE ( TAC )": "TAC", "WAREHOUSE ( WELLGROW - WGR )": "WGR",
        "WAREHOUSE ( PHRA PRADAENG )": "PPD", "WAREHOUSE ( SSW )": "SSW", "WAREHOUSE ( BANGPLEE - WBP )": "BP1",
        "WAREHOUSE ( BANGPLEE - WH.A )": "BPA", "WAREHOUSE ( BANGPLEE - WH.B )": "BPB", "WAREHOUSE ( BANGPLEE - WH.C )": "BPC",
        "WAREHOUSE ( BANGPLEE - WH.D )": "BPD", "WAREHOUSE ( ST9 )": "ST9", "WAREHOUSE ( NEXT GEN )": "NeG",
        "WAREHOUSE ( NeG )": "NeG", "WAREHOUSE ( ST11 )": "ST11", "WAREHOUSE ( STF )": "STF",
        "WAREHOUSE ( TBN )": "TBN1", "WAREHOUSE (P304 )": "P304", "WAREHOUSE (UAF )": "UAF",
        "WAREHOUSE ( HMW )": "HMW", "WAREHOUSE ( TBN2 )": "TBN2", "WAREHOUSE (BN20 )": "BN20",
        "WAREHOUSE (OTHER)": "OTHER", "WAREHOUSE (ZC )": "ZC"
    }

    def extract_site(desc):
        return warehouse_site_map.get(desc.strip(), "Unknown")

    all_data = []
    for file in files:
        df = pd.read_excel(file, sheet_name=0)
        df_str = df.astype(str).apply(lambda x: x.str.upper(), axis=1)

        customer_row = df_str[df_str.eq("CUSTOMER ").any(axis=1)].iloc[0]
        customer_col = customer_row[customer_row.str.contains("CUSTOMER", regex=True)].index[0]

        revenue_row = df_str[df_str.eq("REVENUE").any(axis=1)].iloc[0]
        amount_col = revenue_row[revenue_row == "REVENUE"].index[0]

        warehouse_rows = df[df.iloc[:, 0].astype(str).str.contains("WAREHOUSE", case=False, na=False)].index
        for i, row_index in enumerate(warehouse_rows):
            site = extract_site(df.iloc[row_index, 0])
            next_row = warehouse_rows[i+1] if i+1 < len(warehouse_rows) else len(df)
            df.loc[row_index+1:next_row-1, 'Site'] = site

        data = df[[customer_col, amount_col, 'Site']].dropna()
        data = data[~data.iloc[:, 0].astype(str).str.contains("Customer", case=False)]

        date = datetime.strptime(file.name.split('.')[0], "%Y%m")
        data['Date'] = date.replace(day=1)
        data['Month'] = date.month
        data['Year'] = date.year
        data.columns = ['Customer', 'Amount', 'Site', 'Date', 'Month', 'Year']
        all_data.append(data)

    df_final = pd.concat(all_data, ignore_index=True)

    df_all_site = df_final.groupby(['Year', 'Month'], as_index=False).agg({'Date': 'first', 'Amount': 'sum'})
    df_all_site['Site'] = 'SDCT'
    df_all_site['Customer'] = 'All Customer'

    df_final = pd.concat([df_final, df_all_site], ignore_index=True)

    custom_order = ['SDCT', 'ACW', 'BPA', 'BPB', 'BPC', 'BPD', 'BN20']
    all_sites = df_final['Site'].unique().tolist()
    full_order = custom_order + [x for x in all_sites if x not in custom_order]
    df_final['Site'] = pd.Categorical(df_final['Site'], categories=full_order, ordered=True)
    df_final = df_final.sort_values(by=['Site', 'Year'])
    df_final['Month'] = df_final['Month'].astype(str).str.zfill(2)

    pivot_df = pd.pivot_table(
        df_final,
        index=['Site', 'Customer', 'Year'],
        columns='Month',
        values='Amount',
        aggfunc='sum',
        fill_value=0,
        observed=False
    )
    pivot_df['Grand Total'] = pivot_df.sum(axis=1)
    pivot_df = pivot_df.reset_index()

    month_map = {f"{i:02d}": calendar.month_abbr[i] for i in range(1, 13)}
    pivot_df = pivot_df.rename(columns=month_map)

    final_columns = ['Site', 'Customer', 'Year'] + list(month_map.values()) + ['Grand Total']
    pivot_df = pivot_df.reindex(columns=final_columns)
    return pivot_df

# --- UI Actions ---
if uploaded_files:
    col1, col2 = st.columns(2)
    with col1:
        if st.button("📊 Generate Official Report"):
            df_official = generate_official_report(uploaded_files)
            st.success("✅ Official Report Ready")
            st.dataframe(df_official.head())
            format_and_download(df_official, report_type="official")

    with col2:
        if st.button("👥 Generate Customer Report"):
            df_customer = generate_customer_report(uploaded_files)
            st.success("✅ Customer Report Ready")
            st.dataframe(df_customer.head())
            format_and_download(df_customer, report_type="customer")

# --- Logout ---
st.markdown("---")
if st.button("🚪 Logout"):
    st.session_state.authenticated = False
    st.rerun()
