# 1_Official_Report.py

import streamlit as st
import pandas as pd
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import io
import plotly.express as px

st.title("📊 Official Monthly Report")

if "show_uploader" not in st.session_state:
    st.session_state.show_uploader = False

#if st.button("📂 Generate Official Report"):
#    st.session_state.show_uploader = True

if st.session_state.show_uploader:
    uploaded_files = st.file_uploader("Upload .xlsx files", type="xlsx", accept_multiple_files=True)

    if uploaded_files:
        st.success("✅ Files uploaded. Generating report...")

        def generate_official_report(files):
            df_list = []
            progress_bar = st.progress(0)
            for i, file in enumerate(files):
                progress_bar.progress(int((i + 1) / len(files) * 100))
                filename = file.name
                month = filename[4:6]
                year = filename[0:4]

                header = pd.read_excel(file, nrows=3, header=None)
                last_col_index = header.iloc[2].tolist().index("Total") if "Total" in header.iloc[2].tolist() else None
                usecols = list(range(last_col_index)) if last_col_index is not None else None

                df = pd.read_excel(file, skiprows=2, nrows=49, usecols=usecols)
                df.columns.values[0:3] = ['Type', 'Item', 'Item Detail']
                df_long = df.melt(id_vars=['Type', 'Item', 'Item Detail'], var_name='Site', value_name='Amount')
                df_long['Month'] = month
                df_long['Year'] = year
                df_list.append(df_long)

            df_final = pd.concat(df_list, ignore_index=True)
            df_final['Amount'] = pd.to_numeric(df_final['Amount'], errors='coerce').fillna(0)
            df_final['Month'] = df_final['Month'].astype(str).str.zfill(2)
            month_map = {f"{i:02d}": calendar.month_abbr[i] for i in range(1, 13)}

            pivot_df = pd.pivot_table(
                df_final,
                index=['Site', 'Item Detail', 'Year'],
                columns='Month',
                values='Amount',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )
            pivot_df['Grand Total'] = pivot_df.sum(axis=1)
            pivot_df = pivot_df.reset_index()
            pivot_df = pivot_df.rename(columns=month_map)

            final_columns = ['Site', 'Item Detail', 'Year'] + list(month_map.values()) + ['Grand Total']
            pivot_df = pivot_df.reindex(columns=final_columns)

            return pivot_df, df_final

        def format_and_download(df_result):
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_result.to_excel(writer, index=False, sheet_name="Report")

            wb = load_workbook(buffer)
            ws = wb["Report"]

            col_index_map = {col: idx + 1 for idx, col in enumerate(df_result.columns)}
            month_names = [calendar.month_abbr[i] for i in range(1, 13)]
            number_cols = month_names + ['Grand Total']

            for col in number_cols:
                col_idx = col_index_map.get(col)
                col_letter = get_column_letter(col_idx)
                for cell in ws[col_letter][1:]:
                    cell.number_format = '#,##0_);[Red](#,##0)'

            if 'Item Detail' in col_index_map:
                ws.column_dimensions[get_column_letter(col_index_map['Item Detail'])].width = 35
            if 'Year' in col_index_map:
                for cell in ws[get_column_letter(col_index_map['Year'])][1:]:
                    cell.alignment = Alignment(horizontal='center')

            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions

            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

            st.download_button(
                label="📥 Download Official Monthly Report",
                data=final_buffer,
                file_name="official_monthly_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        df_result, df_raw = generate_official_report(uploaded_files)
        format_and_download(df_result)

        # ✅ Add Line Chart
        st.markdown("### 📈 Line Chart Summary")
        df_raw['Amount'] = pd.to_numeric(df_raw['Amount'], errors='coerce').fillna(0)
        df_raw['Period'] = df_raw['Year'] + "-" + df_raw['Month']
        df_raw['Period'] = pd.to_datetime(df_raw['Period'], format="%Y-%m")

        site_list = df_raw['Site'].dropna().unique().tolist()
        selected_site = st.selectbox("Select site to display chart", site_list)

        chart_df = df_raw[df_raw['Site'] == selected_site]
        chart_df = chart_df.groupby(['Period'], as_index=False)['Amount'].sum()

        fig = px.line(
            chart_df,
            x='Period',
            y='Amount',
            title=f"Monthly Revenue for Site: {selected_site}",
        )
        fig.update_traces(mode='lines+markers')
        fig.update_layout(xaxis_title="Period", yaxis_title="Amount")

        st.plotly_chart(fig, use_container_width=True)
