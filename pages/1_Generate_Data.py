# 1_Official_Report.py

import streamlit as st
import pandas as pd
import numpy as np
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import io
import plotly.express as px

st.title("📊 Official Monthly Report")

# ✅ แสดง uploader โดยตรง
uploaded_files = st.file_uploader("📤 Upload .xlsx files", type="xlsx", accept_multiple_files=True)

if uploaded_files:
    st.success("✅ Files uploaded. Generating report...")

    def generate_official_report(files):
        df_list = []
        progress_bar = st.progress(0)
        for i, file in enumerate(files):
            progress_bar.progress(int((i + 1) / len(files) * 100))
            filename = file.name
            month = filename[4:6]
            year = filename[0:4]

            header = pd.read_excel(file, nrows=3, header=None)
            last_col_index = header.iloc[2].tolist().index("Total") if "Total" in header.iloc[2].tolist() else None
            usecols = list(range(last_col_index)) if last_col_index is not None else None

            df = pd.read_excel(file, skiprows=2, nrows=49, usecols=usecols)
            df.columns.values[0:3] = ['Type', 'Item', 'Item Detail']
            df_long = df.melt(id_vars=['Type', 'Item', 'Item Detail'], var_name='Site', value_name='Amount')
            df_long['Month'] = month
            df_long['Year'] = year
            df_list.append(df_long)

        df_final = pd.concat(df_list, ignore_index=True)
        df_final['Amount'] = pd.to_numeric(df_final['Amount'], errors='coerce').fillna(0)



        corrections = {
            'Signboard TAX': 'Signboard Tax',
            'Common Expense': 'Common expense',
            'System service': 'Cargo Master',  
            # Add more as needed
        }

        # Apply corrections to 'Item Detail' column
        df_final['Item Detail'] = df_final['Item Detail'].replace(corrections)        

        # Prefix definitions
        special_prefixes = {
            (np.nan, 'Revenue Total'): '[1045]-Revenue Total',
            (np.nan, 'Variable Cost'): '[1046]-Variable Cost',
            (np.nan, 'Marginal Profit'): '[1047]-Marginal Profit',
            (np.nan, 'Fix Cost'): '[1048]-Fix Cost',
            (np.nan, 'Cost Total'): '[1049]-Cost Total',
            (np.nan, 'Gross Profit'): '[1050]-Gross Profit',
            (np.nan, 'Expense Total'): '[1051]-Expense Total',
            (np.nan, 'Operating Profit'): '[1052]-Operate Profit'
        }

        # Filter and create prefix map
        mask = ~df_final.apply(lambda row: (row['Item'], row['Item Detail']) in special_prefixes, axis=1)
        unique_items = df_final[mask][['Item', 'Item Detail']].drop_duplicates().reset_index(drop=True)
        prefix_map = {
            (row['Item'], row['Item Detail']): f"[{str(i+1001).zfill(4)}]-{row['Item Detail']}"
            for i, row in unique_items.iterrows()
        }

        prefix_map.update(special_prefixes)
        
        df_final['Item Detail'] = df_final.apply(
            lambda row: prefix_map.get((row['Item'], row['Item Detail']), row['Item Detail']),
            axis=1
        )

        # VC, FC, MP
        df_vc = df_final[df_final['Type'] == 'v'].groupby(['Site', 'Year', 'Month'], as_index=False)['Amount'].sum()
        df_vc['Item Detail'] = '[1046]-Variable Cost'
        df_vc[['Type', 'Item']] = ''

        df_fc = df_final[df_final['Type'] == 'f'].groupby(['Site', 'Year', 'Month'], as_index=False)['Amount'].sum()
        df_fc['Item Detail'] = '[1048]-Fix Cost'
        df_fc[['Type', 'Item']] = ''
        
        revenue_key = next(((k, v) for k, v in prefix_map.items() if k[1] == 'Revenue'), (None, None))[1]
        df_rev = df_final[df_final['Item Detail'] == revenue_key].groupby(['Site', 'Year', 'Month'], as_index=False)['Amount'].sum()
        df_rev.rename(columns={'Amount': 'Revenue'}, inplace=True)
        df_mp = pd.merge(df_rev, df_vc[['Site', 'Year', 'Month', 'Amount']], on=['Site', 'Year', 'Month'], how='left')
        df_mp.rename(columns={'Amount': 'Variable Cost'}, inplace=True)
        df_mp.fillna(0, inplace=True)
        df_mp['Amount'] = df_mp['Revenue'] - df_mp['Variable Cost']
        df_mp['Item Detail'] = '[1047]-Marginal Profit'
        df_mp[['Type', 'Item']] = ''

        df_mp = df_mp[['Month', 'Year', 'Type', 'Item', 'Item Detail', 'Site', 'Amount']]
        df_vc = df_vc[['Month', 'Year', 'Type', 'Item', 'Item Detail', 'Site', 'Amount']]
        df_fc = df_fc[['Month', 'Year', 'Type', 'Item', 'Item Detail', 'Site', 'Amount']]

        df_final = pd.concat([df_final, df_vc, df_fc, df_mp], ignore_index=True)
        sort_order = {v: i for i, v in enumerate(sorted(prefix_map.values()))}
        df_final['ItemSortOrder'] = df_final['Item Detail'].map(sort_order)
        df_final = df_final.sort_values(by=['Year', 'Month', 'Site', 'Item Detail'])

        df_all_site = (
            df_final.groupby(['Item Detail', 'Year', 'Month'], as_index=False)
            .agg({'Amount': 'sum'})
        )
        df_all_site['Site'] = 'SDCT'

        df_final = pd.concat([df_final, df_all_site], ignore_index=True)

        # กำหนดลำดับหลัก
        custom_order = ['SDCT', 'ACW', 'BPA', 'BPB', 'BPC', 'BPD', 'BN20']

        # ค่าทั้งหมดที่เจอในคอลัมน์ Site
        all_sites = df_final['Site'].unique().tolist()

        # เพิ่มค่าอื่น ๆ ต่อท้ายลำดับ custom
        full_order = custom_order + [x for x in all_sites if x not in custom_order]

        # ทำต่อเหมือนเดิม
        site_order = pd.CategoricalDtype(categories=full_order, ordered=True)
        df_final['Site'] = df_final['Site'].astype(site_order)
        df_final = df_final.sort_values(by=['Site', 'Item Detail', 'Year'])





        
        df_final['Month'] = df_final['Month'].astype(str).str.zfill(2)
        month_map = {f"{i:02d}": calendar.month_abbr[i] for i in range(1, 13)}

        pivot_df = pd.pivot_table(
            df_final,
            index=['Site', 'Item Detail', 'Year'],
            columns='Month',
            values='Amount',
            aggfunc='sum',
            fill_value=0,
            observed=False
        )
        pivot_df['Grand Total'] = pivot_df.sum(axis=1)
        pivot_df = pivot_df.reset_index()
        pivot_df = pivot_df.rename(columns=month_map)

        final_columns = ['Site', 'Item Detail', 'Year'] + list(month_map.values()) + ['Grand Total']
        pivot_df = pivot_df.reindex(columns=final_columns)

        return pivot_df, df_final

    def format_and_download(df_result):
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_result.to_excel(writer, index=False, sheet_name='Report')

        wb = load_workbook(buffer)
        ws = wb["Report"]

        col_index_map = {col: idx + 1 for idx, col in enumerate(df_result.columns)}
        month_names = [calendar.month_abbr[i] for i in range(1, 13)]
        number_cols = month_names + ['Grand Total']

        for col in number_cols:
            col_idx = col_index_map.get(col)
            if col_idx:
                col_letter = get_column_letter(col_idx)
                for cell in ws[col_letter][1:]:
                    cell.number_format = '#,##0_);[Red](#,##0)'

        if 'Item Detail' in col_index_map:
            ws.column_dimensions[get_column_letter(col_index_map['Item Detail'])].width = 35
        if 'Year' in col_index_map:
            for cell in ws[get_column_letter(col_index_map['Year'])][1:]:
                cell.alignment = Alignment(horizontal='center')

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        final_buffer = io.BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)

        st.download_button(
            label="📥 Download Official Monthly Report",
            data=final_buffer,
            file_name="official_monthly_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    df_pivot, df_raw = generate_official_report(uploaded_files)
    format_and_download(df_pivot)

    # 👉 Save to session for use in Chart page
    st.session_state["official_data"] = df_raw
